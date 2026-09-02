import os,re,json,time,tempfile,unicodedata
from urllib.parse import quote_plus,urlparse
import requests
import pandas as pd
import pyarrow.parquet as pq
from huggingface_hub import HfApi
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment

LIMIT=int(os.getenv('UAE_LIMIT','10000'))
MAX_PEOPLE_SHARDS=int(os.getenv('MAX_PEOPLE_SHARDS','1200'))
MAX_PEOPLE_SECONDS=int(os.getenv('MAX_PEOPLE_SECONDS','15000'))
OUT='output'; os.makedirs(OUT,exist_ok=True)
REPO='hugging-science/opendata'
BASE=f'https://huggingface.co/datasets/{REPO}/resolve/main/'
TARGET_RE=re.compile(r'(chief executive|\bceo\b|managing director|general manager|chief marketing|\bcmo\b|marketing director|head of marketing|marketing manager|brand director|head of brand|partnership|sponsorship|commercial director|chief commercial|communications director|head of communications|media director|public relations|\bpr director\b|business development director|head of business development)',re.I)
S=requests.Session(); S.headers.update({'User-Agent':'UAE-commercial-intelligence-research/1.0'})

def norm(s):
    s=unicodedata.normalize('NFKD',str(s or '')).encode('ascii','ignore').decode().lower()
    s=re.sub(r'[^a-z0-9 ]+',' ',s); return re.sub(r'\s+',' ',s).strip()
def domain(w):
    if not w:return ''
    w=str(w).strip(); w=w if w.startswith('http') else 'https://'+w
    try:return urlparse(w).netloc.lower().replace('www.','').split(':')[0]
    except:return ''
def parts(name):
    ws=norm(name).split()
    if len(ws)<2:return '',''
    f=ws[0]; l=''.join(ws[-2:]) if len(ws)>=3 and ws[-2] in {'al','el','bin','ibn','bint'} else ws[-1]
    return f,l
def perms(name,d):
    f,l=parts(name)
    if not f or not l or not d:return ''
    return '; '.join(dict.fromkeys([f'{f}.{l}@{d}',f'{f[0]}{l}@{d}',f'{f[0]}.{l}@{d}',f'{f}{l}@{d}',f'{f}@{d}',f'{l}.{f}@{d}']))
def choose(cols,candidates,contains=()):
    low={c.lower():c for c in cols}
    for x in candidates:
        if x.lower() in low:return low[x.lower()]
    for c in cols:
        if any(x in c.lower() for x in contains):return c
    return None

def download(path,tmp):
    url=BASE+path
    for attempt in range(8):
        r=S.get(url,stream=True,timeout=120,allow_redirects=True)
        if r.status_code==200:
            with open(tmp,'wb') as f:
                for ch in r.iter_content(1024*1024):
                    if ch:f.write(ch)
            return
        if r.status_code in (429,500,502,503,504):
            wait=min(60,3*(2**attempt)); print('retry',r.status_code,path,'wait',wait,flush=True); time.sleep(wait); continue
        r.raise_for_status()
    raise RuntimeError(f'Could not download {path}')

def frame_from_shard(path,needed=None):
    fd,tmp=tempfile.mkstemp(suffix='.parquet'); os.close(fd)
    try:
        download(path,tmp)
        pf=pq.ParquetFile(tmp)
        cols=pf.schema_arrow.names
        use=[c for c in (needed or cols) if c in cols]
        return pq.read_table(tmp,columns=use).to_pandas(),cols
    finally:
        try:os.remove(tmp)
        except:pass

def save_excel(df,path):
    wb=Workbook(); ws=wb.active; ws.title='UAE 10K Master'; ws.append(list(df.columns))
    for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(wrap_text=True)
    for row in df.fillna('').itertuples(index=False,name=None): ws.append(list(row))
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns:
        letter=col[0].column_letter
        ws.column_dimensions[letter].width=min(max(12,max(len(str(x.value or '')) for x in col[:120])+2),42)
    for row in ws.iter_rows():
        for c in row:c.alignment=Alignment(vertical='top',wrap_text=True)
    wb.save(path)

files=HfApi().list_repo_files(REPO,repo_type='dataset')
company_paths=sorted(p for p in files if p.startswith('data/companies/parquet/') and p.endswith('.parquet'))
people_paths=sorted(p for p in files if p.startswith('data/people/parquet/') and p.endswith('.parquet'))
print('shards',len(company_paths),len(people_paths),flush=True)

# ----- COMPANIES: sequential shards, so no thousands of simultaneous HEAD requests -----
companies=[]; company_schema=[]
for i,path in enumerate(company_paths):
    df,cols=frame_from_shard(path)
    company_schema=cols
    c2=choose(cols,['bq_company_address1_country_code2','country_code2','country_code'],('country_code2',))
    c3=choose(cols,['bq_company_address1_country_code3'],('country_code3',))
    if not c2 and not c3: continue
    mask=pd.Series(False,index=df.index)
    if c2: mask|=df[c2].fillna('').astype(str).str.upper().isin(['AE','ARE','UNITED ARAB EMIRATES'])
    if c3: mask|=df[c3].fillna('').astype(str).str.upper().eq('ARE')
    sub=df.loc[mask].copy()
    if len(sub): companies.append(sub)
    n=sum(len(x) for x in companies)
    print(f'company shard {i+1}/{len(company_paths)} UAE={n}',flush=True)
    if n>=LIMIT: break
    time.sleep(.3)
if not companies: raise RuntimeError('No UAE companies found')
raw=pd.concat(companies,ignore_index=True).head(LIMIT)
cols=list(raw.columns)
getc=lambda cand,contains=(): choose(cols,cand,contains)
idc=getc(['bq_id','company_id','id'],('company_id','bq_id')); namec=getc(['bq_company_name','company_name','name'],('company_name',))
webc=getc(['bq_website','website','domain'],('website',)); legalc=getc(['bq_company_legal_name','legal_name'],('legal_name',)); cityc=getc(['bq_company_address1_city','city'],('address1_city',)); statec=getc(['bq_company_address1_state','state','emirate'],('address1_state',)); a1=getc(['bq_company_address1_line_1'],('address1_line_1',)); a2=getc(['bq_company_address1_line_2'],('address1_line_2',)); linkc=getc(['bq_company_linkedin_url','linkedin_url'],('linkedin_url',)); cbc=getc(['bq_company_crunchbase_url','crunchbase_url'],('crunchbase',)); leic=getc(['bq_company_lei','lei'],('company_lei',)); tickc=getc(['bq_ticker','ticker'],('ticker',))
def col(c): return raw[c].fillna('').astype(str) if c else pd.Series(['']*len(raw))
company=pd.DataFrame({'company_id':col(idc),'company_name':col(namec),'legal_name':col(legalc),'website':col(webc),'city':col(cityc),'state_emirate':col(statec),'address':(col(a1)+' '+col(a2)).str.strip(),'linkedin_company':col(linkc),'crunchbase':col(cbc),'lei':col(leic),'ticker':col(tickc)})
company=company[company.company_name.str.strip().ne('')].drop_duplicates(subset=['company_id','company_name']).head(LIMIT).reset_index(drop=True)
company['domain']=company.website.map(domain); company['country']='United Arab Emirates'; company['source']='OpenData Consortium / Hugging Face'; company['source_date']='2026-09-02'
company.to_csv(f'{OUT}/UAE_10K_Companies.csv',index=False)
# Save guaranteed company-only Excel immediately.
base=company.copy();
for c in ['person_name','person_title','published_work_email','business_phone_or_mobile','person_linkedin']:base[c]=''
base['email_permutations']=''; base['email_status']='Not found'; base['meta_ad_library_url']=base.company_name.map(lambda n:'https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=AE&q='+quote_plus(str(n))); base['meta_cohort_note']='Private Meta audiences/cookies are not public; infer cohorts only from public campaigns.'; base['martech_status']='Public website scan pending'; base['validation_status']='No person email tested'
base.to_csv(f'{OUT}/UAE_10K_Master.csv',index=False); save_excel(base,f'{OUT}/UAE_10K_Master.xlsx')
print('base 10K saved',len(company),flush=True)

# ----- PEOPLE: sequential, narrow columns, filter to target company IDs/titles -----
target=set(company.company_id.astype(str)); contacts=[]; start=time.time(); people_schema=[]
for i,path in enumerate(people_paths[:MAX_PEOPLE_SHARDS]):
    if time.time()-start>MAX_PEOPLE_SECONDS:
        print('people time budget reached',flush=True); break
    # Discover schema from first shard, then request only useful fields in subsequent shards.
    if not people_schema:
        sdf,people_schema=frame_from_shard(path)
        pcid=choose(people_schema,['bq_company_id','company_id','employer_company_id','current_company_id','bq_employer_id'],('company_id','employer_id'))
        ptitle=choose(people_schema,['bq_job_title','job_title','title','position','current_title'],('job_title','current_title'))
        pname=choose(people_schema,['bq_full_name','full_name','name','person_name'],('full_name','person_name')); pfirst=choose(people_schema,['bq_first_name','first_name'],('first_name',)); plast=choose(people_schema,['bq_last_name','last_name'],('last_name',)); pemail=choose(people_schema,['bq_work_email','work_email','email','business_email'],('work_email','business_email')); pphone=choose(people_schema,['bq_phone','phone','mobile_phone','work_phone'],('phone',)); plink=choose(people_schema,['bq_linkedin_url','linkedin_url','person_linkedin_url'],('linkedin_url',))
        needed=list(dict.fromkeys(x for x in [pcid,ptitle,pname,pfirst,plast,pemail,pphone,plink] if x)); df=sdf[needed] if needed else sdf
        if not pcid or not ptitle: print('people schema has no join/title',people_schema,flush=True); break
    else:
        needed=list(dict.fromkeys(x for x in [pcid,ptitle,pname,pfirst,plast,pemail,pphone,plink] if x)); df,_=frame_from_shard(path,needed)
    cid=df[pcid].fillna('').astype(str); titles=df[ptitle].fillna('').astype(str)
    m=cid.isin(target)&titles.map(lambda x:bool(TARGET_RE.search(x)))
    if m.any():
        d=df.loc[m].copy(); out=pd.DataFrame({'company_id':d[pcid].fillna('').astype(str),'person_title':d[ptitle].fillna('').astype(str)})
        if pname: out['person_name']=d[pname].fillna('').astype(str)
        else: out['person_name']=(d[pfirst].fillna('').astype(str) if pfirst else '')+' '+(d[plast].fillna('').astype(str) if plast else '')
        out['published_work_email']=d[pemail].fillna('').astype(str) if pemail else ''
        out['business_phone_or_mobile']=d[pphone].fillna('').astype(str) if pphone else ''
        out['person_linkedin']=d[plink].fillna('').astype(str) if plink else ''
        contacts.append(out)
    if (i+1)%10==0: print(f'people shard {i+1}/{min(MAX_PEOPLE_SHARDS,len(people_paths))} contacts={sum(len(x) for x in contacts)}',flush=True)
    time.sleep(.25)

if contacts:
    ppl=pd.concat(contacts,ignore_index=True).drop_duplicates(subset=['company_id','person_name','person_title'])
    ppl['priority']=ppl.person_title.str.lower().map(lambda t:0 if ('chief marketing' in t or 'cmo' in t or 'head of marketing' in t or 'marketing director' in t) else (1 if ('chief executive' in t or 'ceo' in t or 'managing director' in t) else 2))
    ppl=ppl.sort_values(['company_id','priority']).groupby('company_id',as_index=False,group_keys=False).head(8).drop(columns='priority')
    master=company.merge(ppl,on='company_id',how='left')
else: master=base.copy()
for c in ['person_name','person_title','published_work_email','business_phone_or_mobile','person_linkedin']:
    if c not in master:master[c]=''
master['email_permutations']=master.apply(lambda r:perms(r.get('person_name',''),r.get('domain','')),axis=1)
master['email_status']=master.apply(lambda r:'Published' if str(r.get('published_work_email','')).strip() not in ('','nan','None') else ('Candidate only' if str(r.get('person_name','')).strip() not in ('','nan','None') and r.get('domain') else 'Not found'),axis=1)
master['meta_ad_library_url']=master.company_name.map(lambda n:'https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=AE&q='+quote_plus(str(n)))
master['meta_cohort_note']='Private Meta audiences/cookies are not public; infer cohorts only from public campaigns.'
master['martech_status']='Public website scan pending'
master['validation_status']=master.email_status.map(lambda x:'Published source record; mailbox verification still required' if x=='Published' else ('Pattern candidate; NOT mailbox-verified' if x=='Candidate only' else 'No email'))
master.to_csv(f'{OUT}/UAE_10K_Master.csv',index=False); save_excel(master,f'{OUT}/UAE_10K_Master.xlsx')
summary={'companies':int(len(company)),'master_rows':int(len(master)),'contacts_found':int(sum(len(x) for x in contacts)) if contacts else 0,'company_shards_used':i+1 if 'i' in locals() else 0,'people_shards_scanned':min((i+1 if 'i' in locals() else 0),MAX_PEOPLE_SHARDS),'people_coverage_companies':int(master.person_name.fillna('').astype(str).ne('').groupby(master.company_id).max().sum()) if 'person_name' in master else 0,'file':'UAE_10K_Master.xlsx','note':'Email permutations are candidates; verification status is explicit. Meta private audiences/cookies are not collected.'}
with open(f'{OUT}/summary.json','w') as f:json.dump(summary,f,indent=2)
print(json.dumps(summary,indent=2),flush=True)
