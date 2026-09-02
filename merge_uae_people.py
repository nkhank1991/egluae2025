import os,csv,json,glob,re
from collections import Counter
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

OUT='output'; os.makedirs(OUT,exist_ok=True)
people_files=sorted(glob.glob('shards/**/UAE_people_shard_*.csv',recursive=True)+glob.glob('output/UAE_people_shard_*.csv'))
audit_files=sorted(glob.glob('shards/**/UAE_audit_shard_*.csv',recursive=True)+glob.glob('output/UAE_audit_shard_*.csv'))
summary_files=sorted(glob.glob('shards/**/summary_*.json',recursive=True)+glob.glob('output/summary_*.json'))
if not people_files: raise SystemExit('No people shard files found')

def read_csvs(paths):
 rows=[]
 for p in paths:
  with open(p,encoding='utf-8-sig',errors='replace',newline='') as f: rows.extend(csv.DictReader(f))
 return rows

people=read_csvs(people_files)
# Deduplicate same person/title/domain discovered from multiple place records/pages.
uniq={}
for r in people:
 k=(r.get('domain','').lower().strip(),r.get('person_name','').lower().strip(),r.get('canonical_role','').lower().strip())
 if not all(k): continue
 old=uniq.get(k)
 if old is None: uniq[k]=r
 else:
  # prefer a row with published person email, then richer source pages.
  score=lambda x:(1 if x.get('published_person_email') else 0, int(x.get('pattern_evidence_count') or 0), len(x.get('pages_crawled','')))
  if score(r)>score(old): uniq[k]=r
people=list(uniq.values())
# Sort marketing/chief contacts first.
rank={'Marketing':0,'Brand':1,'Partnerships/Sponsorship':2,'Commercial/Growth':3,'Communications/Media/PR':4,'Executive':5,'Other':6}
people.sort(key=lambda r:(r.get('company_name','').lower(),rank.get(r.get('role_family','Other'),9),r.get('person_name','').lower()))

fields=['company_name','category','website','domain','person_name','current_title','canonical_role','role_family','currentness_basis','person_source_url','research_date','published_person_email','published_company_emails','observed_email_pattern','pattern_evidence_count','preferred_inferred_email','email_permutation_candidates','email_source_type','email_confidence','domain_has_mx','mailbox_verification_status','public_business_phones','social_urls','pages_crawled','base_source']
with open(f'{OUT}/UAE_Current_Chief_Marketing_Contacts.csv','w',newline='',encoding='utf-8-sig') as f:
 w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(people)

def safe(v):
 if v is None:return ''
 if isinstance(v,str):return ILLEGAL_CHARACTERS_RE.sub('',v)[:32767]
 return v
wb=Workbook(write_only=True)
ws=wb.create_sheet('Decision Makers')
header=[]
for name in fields:
 c=WriteOnlyCell(ws,value=name);c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='1F4E78');c.alignment=Alignment(wrap_text=True,vertical='top');header.append(c)
ws.append(header)
for i in range(1,len(fields)+1): ws.column_dimensions[get_column_letter(i)].width=28
for idx in [1,5,6,10,12,16,17,23,24]:
 if idx<=len(fields):ws.column_dimensions[get_column_letter(idx)].width=42
for r in people:
 row=[]
 for name in fields:
  c=WriteOnlyCell(ws,value=safe(r.get(name,'')));c.alignment=Alignment(vertical='top',wrap_text=True);row.append(c)
 ws.append(row)
# Summary sheet
ss=wb.create_sheet('Summary')
counts=Counter(r.get('role_family','Other') for r in people)
published=sum(bool(r.get('published_person_email')) for r in people)
inferred=sum(bool(r.get('preferred_inferred_email')) for r in people)
summary=[['Metric','Value'],['Unique decision makers',len(people)],['Published person emails',published],['Preferred inferred emails',inferred],['Domains represented',len(set(r.get('domain','') for r in people if r.get('domain')))],['Important','Inferred/permutation emails are candidates only and are NOT mailbox-verified.']]
for k,v in sorted(counts.items()):summary.append([k,v])
for rr in summary:ss.append(rr)
wb.save(f'{OUT}/UAE_Current_Chief_Marketing_Contacts.xlsx')

shard_summaries=[]
for p in summary_files:
 try:
  with open(p) as f:shard_summaries.append(json.load(f))
 except:pass
final={'unique_decision_makers':len(people),'published_person_emails':published,'preferred_inferred_emails':inferred,'domains_represented':len(set(r.get('domain','') for r in people if r.get('domain'))),'role_families':dict(counts),'shards':shard_summaries,'note':'Currentness basis is presence on the company website at research time. Candidate/inferred emails are not mailbox-verified.'}
with open(f'{OUT}/people_summary.json','w') as f:json.dump(final,f,indent=2)
print(json.dumps(final,indent=2))
